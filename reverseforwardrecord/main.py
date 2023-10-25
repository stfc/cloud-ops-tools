import csv
from functools import reduce

# for reverse records

'''
 8 and 10 for the tt csv grab parts and 8 is fqdn so goes first when putting them together

have to remove something after the third dot as this is a specific in an ip case where there is only allowed 4 dots in 
ip usually you cant do this everytime but that is how it must be done in this instance, learning how to do this will
also help with removing the fqdn things that arent necessary when reversing

example is 
ip_addr = "17.0.77.25"
from functools import reduce 
out = reduce(lambda x, y: f"{x}.{y}",
ip_addr.split(".")[::-1])
print(out)


WAY TO CONVERT XLSX FILES TO CSV AS I ACCIDENTALLY REMOVED THEM FROM THE PROGRAM
# Convert all sheets
from openpyxl import load_workbook
import csv
 
# Load in the workbook
workbook = load_workbook(filename="employees.xlsx")
# get all sheets on the XLSX file.
sheets = workbook.sheetnames
 
# Loop through each sheet and convert it to CSV
for sheet in sheets:
    # Create the CSV file and write the rows to it
    with open(f"{sheet}.csv", "w", newline="") as infile:
        c = csv.writer(infile)
        # Loop all rows of a given sheet.
        for row in workbook[sheet]:
            c.writerow([cell.value for cell in row])
            

'''


def write_file(values):
    with open('reverse.txt', 'w', encoding="utf-8") as out_file:
        for fqdn, idrac in zip(hypervisors, ip_addresses):
            out_file.write(f"{idrac}    IN PTR    {fqdn+'.'}\n")


# opens file and grabs a certain column
if __name__ == "__main__":

    hypervisors = []
    ip_addresses = []
    with open("tt.csv") as file_handle:
        reader = csv.reader(file_handle)

        column_to_print = 8
        column_to_print2 = 10
        for row in reader:
            fqdn = row[column_to_print]
            idrac = row[column_to_print2]
            hypervisors.append(fqdn)
            idrac = reduce(lambda x, y: f"{x}.{y}", idrac.split(".")[::-1])#
            idrac = idrac.rsplit('.', 1)[0]
            ip_addresses.append(idrac)
            print(fqdn)
            print(idrac)
        write_file(hypervisors)

'''
make an input that checks the dns server inputted from the list made 
dnsresolver python plugin
do this in order to verify if dnslookup works with them 
do this with both forward ip and reverse ip
either batch resolve or do individual input resolve
'''