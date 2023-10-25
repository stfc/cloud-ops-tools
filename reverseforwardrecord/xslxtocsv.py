from openpyxl import load_workbook
import csv

# Load in the workbook
workbook = load_workbook(filename="test.xlsx")
# get all sheets on the XLSX file.
sheets = workbook.sheetnames

# Loop through each sheet and convert it to CSV
for sheet in sheets:
    # Create the CSV file and write the rows to it
    with open(f"{sheet}.csv", "w", newline="") as infile:
        c = csv.writer(infile)
        # Loop all rows of a given sheet.
        for row in workbook[sheet]:
            c.writerow([cell.value for cell in row])
