from typing import List, Dict
from openpyxl import load_workbook
from functools import reduce
import socket
import pandas as pd

"""
read from 2 txt files one containing dns name and another containing ip address
convert into list of dictionaries containing them as keys - dns name and ip address
write a function that takes 2 lists and converts them into a list of dicts
just have to change read from netbox
only converts from xlsx
xlsx file has to have fqdn in column 8 and idrac_ip in column 10
otherwise have to set it up with manual user input
"""


def create_file():
    """
    makes a txt file for the results to be returned into
    :return: output.txt
    """
    f = open("ignored/output.txt", "w")
    f.close()


def read_from_netbox(
        netbox_filepath: str, fqdn_column: int = 8, idrac_ip_column: int = 10
) -> List[Dict]:
    """
    reads from netbox file, in a xlsx format
    :param netbox_filepath: path for the netbox info
    :param fqdn_column: full qualified domain name column in netbox file
    :param idrac_ip_column: idrac ip column in netbox file
    :return: a list of dictionaries containing fqdn and idrac_ip pairs
    """

    workbook = load_workbook(filename=netbox_filepath)
    # get all sheets on the XLSX file.
    sheets = workbook.sheetnames
    workbook_info = []

    for i, sheet in enumerate(sheets, 1):
        for j, row in enumerate(workbook[sheet]):
            if j == 0:
                # ignore headers
                continue

            try:
                val = {
                    "fqdn": row[fqdn_column].value,
                    "idrac_ip": row[idrac_ip_column].value,
                }
            except IndexError as exp:
                raise RuntimeError(
                    f"Error: you've provided an excel sheet: (no. {i}) which does not contain a value in "
                    f"either the fqdn column {fqdn_column} or idrac_ip column {idrac_ip_column}"
                ) from exp
            workbook_info.append(val)
    return workbook_info


def parse_netbox_info(
        netbox_info: List[Dict], reverse_order: bool = False
) -> List[Dict]:
    """
    Parse info from netbox and create corresponding hypervisor fqdn and ip value
    :param netbox_info: list of dictionaries holding fqdn and idrac_ip
    :param reverse_order: use reverse or forward order / if true use forward if false use reverse
    :return: list of dictionaries which will contain hypervisor fqdn and ip keys-value pairs
    """
    parsed_info = []
    for item in netbox_info:
        parsed_item = {}
        if reverse_order:
            idrac = reduce(
                lambda x, y: f"{x}.{y}", item["idrac_ip"].split(".")[::-1]
            )  #
            idrac = idrac.rsplit(".", 1)[0]
            parsed_item["ip_address"] = idrac
            parsed_item["hypervisor"] = item["fqdn"]
        else:
            parsed_item["ip_address"] = item["idrac_ip"]
            parsed_item["hypervisor"] = item["fqdn"].split(".")[0]
        parsed_info.append(parsed_item)
    return parsed_info


def write_output(
        parsed_info: List[Dict], output_filepath: str, reverse_order: bool = False
):
    """
    write parsed netbox info to a file
    :param parsed_info: list of dictionaries which contains hypervisor fqdn and ip keys
    :param output_filepath: file path for output
    :param reverse_order: use reverse or forward order / if true use forward if false use reverse
    :return: the text written into the file
    """
    with open(output_filepath, "w", encoding="utf-8") as out_file:
        for item in parsed_info:
            if reverse_order:
                out_file.write(f"{item['ip_address']}\tIN PTR\t{item['hypervisor']}\n")
            else:
                out_file.write(f"{item['hypervisor']}\tIN A\t{item['ip_address']}\n")


def main(
        netbox_filepath: str,
        fqdn_column: int = 8,
        idrac_ip_column: int = 10,
        reverse_order: bool = False,
        output_filepath: str = "output.txt",
):
    """
    this function will create a file for the output for dns records
    :param netbox_filepath: path for the netbox info
    :param fqdn_column: full qualified domain name column in netbox file
    :param idrac_ip_column: idrac ip column in netbox file
    :param reverse_order: use reverse or forward order / if true use forward if false use reverse
    :param output_filepath: file path for output
    :return: None
    """
    create_file()
    netbox_info = read_from_netbox(netbox_filepath, fqdn_column, idrac_ip_column)
    parsed_netbox_info = parse_netbox_info(netbox_info, reverse_order)
    write_output(parsed_netbox_info, output_filepath, reverse_order)


if __name__ == "__main__":
    netbox_filepath = "ignored/test.xlsx"
    main(netbox_filepath, reverse_order=False)
